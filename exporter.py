from parser import *
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import io

def export_result_to_excel(results, output_stream=None):
    import pandas as pd
    from tempfile import NamedTemporaryFile

    summaries = []
    for result in results:
        summary = {
            "CAS Index Name": result.get("chemical_name"),
            "CAS Registry Number": result.get("cas_number"),
            "Hazard Statements\n(H302, 250 = highlight yellow)\n- Section 2 of the SDS": "\n".join(
                f'{entry.get("ghs_code", "")}: {entry.get("original_text", "")}' for entry in result.get("ghs_from_sds", [])
                if entry.get("ghs_code")
            ),
            "FP < 100 def F\nhighlight yellow -\nSection 9 of the SDS": result.get("flash_point"),
            "Storage statement\n(Store < 4 deg C) -\nSection 7 of the SDS": result.get("storage_conditions"),
            "Reactivity (section 5&10 of the SDS\n- peroxides and HBr, HCl, HCN gas\ngenerated- highlight yellow)": result.get("reactivity"),
            "Physical State": result.get("appearance"),
            "Odor": result.get("odor"),
        }

        nfpa_data = result.get("nfpa", {})
        if nfpa_data:
            if isinstance(nfpa_data, list):
                for idx, nfpa_entry in enumerate(nfpa_data):
                    nfpa_summary = []
                    for key in ["Health", "Flammability", "Instability", "Special"]:
                        value = nfpa_entry.get(key, {}).get("value_html", "N/A")
                        desc = nfpa_entry.get(key, {}).get("description", "").strip()
                        if value is None:
                            value = "N/A"
                        if desc:
                            nfpa_summary.append(f"{key}: {value} ({desc})")
                        else:
                            nfpa_summary.append(f"{key}: {value}")
                    name = nfpa_entry.get("name", "NFPA")
                    nfpa_summary.insert(0, f"Name: {name}")
                    summary[f"NFPA Rating {idx+1}"] = "\n".join(nfpa_summary)
            elif isinstance(nfpa_data, dict):
                nfpa_summary = []
                for key in ["Health", "Flammability", "Instability", "Special"]:
                    value = nfpa_data.get(key, {}).get("value_html", "N/A")
                    desc = nfpa_data.get(key, {}).get("description", "").strip()
                    if value is None:
                        value = "N/A"
                    if desc:
                        nfpa_summary.append(f"{key}: {value} ({desc})")
                    else:
                        nfpa_summary.append(f"{key}: {value}")
                name = nfpa_data.get("name", "NFPA")
                nfpa_summary.insert(0, f"Name: {name}")
                summary["NFPA Rating"] = "\n".join(nfpa_summary)

        summaries.append(summary)

    df = pd.DataFrame(summaries)

    if output_stream is None:
        output_stream = io.BytesIO()

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        df.to_excel(tmp.name, index=False)
        wb = load_workbook(tmp.name)
        ws = wb.active

        # Enable wrap text for all non-header cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str):
                    cell.alignment = Alignment(wrap_text=True)

        # Enable wrap text for cells with newline characters
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and '\n' in cell.value:
                    cell.alignment = Alignment(wrap_text=True)

        # Auto-adjust column widths based on max line length in header cell
        for column_cells in ws.iter_cols(min_row=1, max_row=1):
            for cell in column_cells:
                if cell.value:
                    header_lines = str(cell.value).split('\n')
                    max_length = max(len(line) for line in header_lines)
                    ws.column_dimensions[get_column_letter(cell.column)].width = max_length + 2

        wb.save(output_stream)

    output_stream.seek(0)
    return output_stream